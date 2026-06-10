from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

wb = Workbook()

# ── Sheet 1: 병원 입력 시트 ──
ws = wb.active
ws.title = '병원목록'

thin = Side(style='thin', color='CBD5E1')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 컬럼 정의 (DB컬럼명, 표시명, 너비, 필수여부, 설명)
columns = [
    ('name',          '병원명 *',         28, True,  '필수 | 병원 공식 명칭'),
    ('region',        '지역',              14, False, '예) 서울, 부산, 경기'),
    ('address',       '주소',              36, False, '도로명 주소'),
    ('phone',         '병원 대표전화',      18, False, '예) 02-1234-5678'),
    ('contact_name',  '담당자명(원장명)',   18, False, '원장 또는 담당 직원 이름'),
    ('contact_phone', '담당자 연락처',      18, False, '예) 010-1234-5678'),
    ('contact_email', '담당자 이메일',      26, False, '예) doctor@hospital.com'),
    ('notes',         '메모',              36, False, '납품 히스토리, 특이사항 등'),
]

# ── 안내 행 (1행) ──
ws.merge_cells('A1:H1')
notice = ws.cell(row=1, column=1,
    value='✅ 노란색(*) 필수 항목만 채워도 업로드 가능합니다. id/created_at/token 컬럼은 Supabase가 자동 생성하므로 입력 불필요.')
notice.font = Font(name='Arial', size=9, color='92400E')
notice.fill = PatternFill('solid', start_color='FEF3C7', end_color='FEF3C7')
notice.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 24

# ── DB 컬럼명 행 (2행, 회색 소자) ──
for col_idx, (db_col, _, _, _, _) in enumerate(columns, 1):
    cell = ws.cell(row=2, column=col_idx, value=db_col)
    cell.font = Font(name='Arial', size=8, color='94A3B8', italic=True)
    cell.fill = PatternFill('solid', start_color='F1F5F9', end_color='F1F5F9')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
ws.row_dimensions[2].height = 16

# ── 헤더 행 (3행) ──
for col_idx, (_, label, width, required, _) in enumerate(columns, 1):
    cell = ws.cell(row=3, column=col_idx, value=label)
    if required:
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', start_color='1E3A5F', end_color='1E3A5F')
    else:
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', start_color='334155', end_color='334155')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[3].height = 28

# ── 설명 행 (4행) ──
for col_idx, (_, _, _, required, desc) in enumerate(columns, 1):
    cell = ws.cell(row=4, column=col_idx, value=desc)
    cell.font = Font(name='Arial', size=8, color='64748B', italic=True)
    cell.fill = PatternFill('solid', start_color='F8FAFC', end_color='F8FAFC')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
ws.row_dimensions[4].height = 18

# ── 예시 행 (5행) ──
sample = [
    '○○○의원',
    '서울',
    '서울특별시 강남구 테헤란로 123',
    '02-1234-5678',
    '홍길동',
    '010-9876-5432',
    'hong@clinic.com',
    '2023년 납품 병원 (초음파, 혈압계)',
]
for col_idx, val in enumerate(sample, 1):
    cell = ws.cell(row=5, column=col_idx, value=val)
    cell.font = Font(name='Arial', size=9, color='0369A1', italic=True)
    cell.fill = PatternFill('solid', start_color='EFF6FF', end_color='EFF6FF')
    cell.alignment = Alignment(vertical='center')
    cell.border = border
ws.row_dimensions[5].height = 20

# ── 데이터 입력 영역 (6행~205행, 200개) ──
row_fill_even = PatternFill('solid', start_color='F8FAFC', end_color='F8FAFC')
data_font = Font(name='Arial', size=10)
for row_idx in range(6, 206):
    fill = row_fill_even if row_idx % 2 == 0 else None
    for col_idx in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical='center')
        if fill:
            cell.fill = fill
    ws.row_dimensions[row_idx].height = 20

# 병원명 컬럼 필수 강조 (노란 배경)
for row_idx in range(6, 206):
    cell = ws.cell(row=row_idx, column=1)
    if not cell.value:
        cell.fill = PatternFill('solid', start_color='FFFBEB', end_color='FFFBEB')

# 틀 고정 (3행 헤더까지 고정)
ws.freeze_panes = 'A6'

# 자동 필터
ws.auto_filter.ref = 'A3:H205'


# ── Sheet 2: 업로드 가이드 ──
guide = wb.create_sheet('📋 업로드 가이드')
guide.column_dimensions['A'].width = 16
guide.column_dimensions['B'].width = 70

guide_data = [
    ('', ''),
    ('  📌 업로드 방법', ''),
    ('', ''),
    ('  방법 1', 'Supabase 대시보드 → Table Editor → hospitals → Insert rows → CSV 업로드'),
    ('  방법 2', 'SQL Editor에서 INSERT 구문 직접 실행 (개발자가 변환해줌)'),
    ('  방법 3', '앱에서 일괄 Import 기능 사용 (개발 예정)'),
    ('', ''),
    ('  ⚠️ 주의사항', ''),
    ('', ''),
    ('  입력 불필요', 'id, created_at, token, access_pin → Supabase 자동 생성'),
    ('  필수 항목', 'name(병원명) 만 필수, 나머지는 빈칸 가능'),
    ('  중복 방지', '같은 병원명이 중복되지 않게 확인 필요'),
    ('  전화번호 형식', '하이픈(-) 포함 또는 미포함 모두 가능 (예: 02-1234-5678)'),
    ('', ''),
    ('  📊 DB 컬럼 매핑', ''),
    ('', ''),
    ('  병원명', '→ name'),
    ('  지역', '→ region'),
    ('  주소', '→ address'),
    ('  병원 대표전화', '→ phone'),
    ('  담당자명(원장명)', '→ contact_name'),
    ('  담당자 연락처', '→ contact_phone'),
    ('  담당자 이메일', '→ contact_email'),
    ('  메모', '→ notes'),
]

for row_idx, (col_a, col_b) in enumerate(guide_data, 1):
    ca = guide.cell(row=row_idx, column=1, value=col_a)
    cb = guide.cell(row=row_idx, column=2, value=col_b)
    ca.font = Font(name='Arial', size=10,
                   bold=('📌' in col_a or '⚠️' in col_a or '📊' in col_a))
    cb.font = Font(name='Arial', size=10)
    if '📌' in col_a or '⚠️' in col_a or '📊' in col_a:
        ca.fill = PatternFill('solid', start_color='1E293B', end_color='1E293B')
        ca.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        guide.merge_cells(f'A{row_idx}:B{row_idx}')
        guide.cell(row=row_idx, column=1).value = col_a
    guide.row_dimensions[row_idx].height = 20

out_path = os.path.join('C:/Users/jojo/mediquote', '병원목록_입력양식.xlsx')
wb.save(out_path)
print(f'저장 완료: {out_path}')
