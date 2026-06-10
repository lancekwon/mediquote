from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

wb = Workbook()

COLOR_HEADER   = "1E3A5F"
COLOR_REQUIRED = "DBEAFE"
COLOR_OPTIONAL = "F8FAFC"
COLOR_SPEC     = "EFF6FF"
COLOR_SAMPLE   = "FFFBEB"

thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 시트1: 장비 입력 ──
ws = wb.active
ws.title = "장비 입력"

ws.merge_cells("A1:U1")
ws["A1"] = "DWmedi 장비 정보 입력 양식  |  파란셀=필수 / 흰셀=선택  |  카테고리명은 '카테고리 목록' 시트 참고  |  스펙은 최대 5개"
ws["A1"].font = Font(name="Arial", size=10, bold=True, color="1E3A5F")
ws["A1"].fill = PatternFill("solid", start_color="DBEAFE")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 26

headers = [
    ("카테고리명*",  True),
    ("품목명*",      True),
    ("모델명*",      True),
    ("제조사",       False),
    ("가격(원)",     False),
    ("원산지",       False),
    ("인증",         False),
    ("A/S기간",      False),
    ("보증기간",     False),
    ("제품소개",     False),
    ("스펙1항목",    False),
    ("스펙1값",      False),
    ("스펙2항목",    False),
    ("스펙2값",      False),
    ("스펙3항목",    False),
    ("스펙3값",      False),
    ("스펙4항목",    False),
    ("스펙4값",      False),
    ("스펙5항목",    False),
    ("스펙5값",      False),
    ("특이사항",     False),
]

tips = [
    "'카테고리 목록' 시트에서 확인\n드롭다운 선택 가능",
    "예: X-Ray (DR System)",
    "예: Xvision HF 525R",
    "예: Gemss healthcare",
    "숫자만. 예: 15000000\n미확정시 빈칸",
    "예: 대한민국 / 미국",
    "쉼표 구분\n예: 식약처 허가, CE",
    "예: 2년",
    "예: 1년",
    "200자 이내 핵심 특징",
    "예: 관전압","예: 40~150 kVp",
    "예: 관전류","예: 100~500 mA",
    "","","","","","",
    "예: 콘덴서 포함",
]

ws.row_dimensions[2].height = 36
for i, (hdr, req) in enumerate(headers, 1):
    c = ws.cell(row=2, column=i, value=hdr)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    tip = tips[i-1] if i-1 < len(tips) else ""
    if tip:
        cm = Comment(tip, "DWmedi")
        cm.width = 180
        cm.height = 55
        c.comment = cm

ws.row_dimensions[3].height = 16
for i, (_, req) in enumerate(headers, 1):
    c = ws.cell(row=3, column=i)
    label = "필수" if req else ("스펙" if 11 <= i <= 20 else "선택")
    bg = "BFDBFE" if label == "필수" else ("E0EDFF" if label == "스펙" else "F1F5F9")
    c.value = label
    c.font = Font(name="Arial", size=8, color="475569")
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

samples = [
    ["영상진단장비","X-Ray (DR System)","Xvision HF 525R","Gemss healthcare",15000000,
     "대한민국","식약처 허가, CE, ISO 13485","2년","1년",
     "고주파 DR X-Ray 시스템. 의원급·병원급 모두 적합.",
     "관전압","40~150 kVp","관전류","100~500 mA","촬영방식","DR (FPD)","해상도","3.4 lp/mm","","",""],
    ["물리치료재활장비","ICT 간섭파","Ecoplus","Goodple",1000000,
     "대한민국","식약처 허가","2년","1년","합리적 가격의 ICT 간섭파 치료기.",
     "채널","4채널","주파수","1~150 Hz","","","","","","",""],
    ["영상진단장비","C-Arm","KMC-650","Gemss healthcare",34500000,
     "대한민국","식약처 허가, CE","2년","1년","정형외과·척추외과 전용 C-Arm.",
     "관전압","40~110 kVp","이미지 증배관","9인치 II","회전범위","C자 190도","","","","",""],
]

for r, row in enumerate(samples, 4):
    ws.row_dimensions[r].height = 20
    for c_idx, val in enumerate(row, 1):
        c = ws.cell(row=r, column=c_idx, value=val)
        c.font = Font(name="Arial", size=9, color="374151")
        c.fill = PatternFill("solid", start_color=COLOR_SAMPLE)
        c.alignment = Alignment(vertical="center")
        c.border = border

for r in range(7, 207):
    ws.row_dimensions[r].height = 20
    for c_idx, (_, req) in enumerate(headers, 1):
        c = ws.cell(row=r, column=c_idx)
        bg = COLOR_REQUIRED if req else (COLOR_SPEC if 11 <= c_idx <= 20 else COLOR_OPTIONAL)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = border
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(vertical="center")

widths = [18,20,22,20,14,12,22,10,10,36,12,18,12,18,12,18,12,18,12,18,22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A4"

cats = [
    "영상진단장비","물리치료재활장비","기타의료기기","수술장비","재활치료장비",
    "소독멸균장비","진단검사장비","치과장비","한방장비","안과장비",
    "피부미용장비","수술기구소모품","기타"
]
dv = DataValidation(type="list", formula1='"' + ",".join(cats) + '"', allow_blank=True)
dv.sqref = "A4:A206"
ws.add_data_validation(dv)

# ── 시트2: 카테고리 목록 ──
ws2 = wb.create_sheet("카테고리 목록")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 42

ws2.merge_cells("A1:B1")
ws2["A1"] = "카테고리 목록  —  장비 입력 시 아래 카테고리명을 정확히 복사해서 사용하세요"
ws2["A1"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", start_color=COLOR_HEADER)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for i, h in enumerate(["카테고리명", "설명 (참고용)"], 1):
    c = ws2.cell(row=2, column=i, value=h)
    c.font = Font(name="Arial", size=10, bold=True, color="1E3A5F")
    c.fill = PatternFill("solid", start_color="DBEAFE")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
    ws2.row_dimensions[2].height = 22

cat_desc = [
    ("영상진단장비",    "X-Ray, C-Arm, 초음파, 골밀도, PACS 등"),
    ("물리치료재활장비","ICT, Microwave, 레이저, ESWT, 크라이오 등"),
    ("기타의료기기",    "분류가 명확하지 않은 의료기기"),
    ("수술장비",        "수술용 장비 및 관련 기구"),
    ("재활치료장비",    "재활 전용 치료 장비"),
    ("소독멸균장비",    "소독기, 멸균기, UV 살균기 등"),
    ("진단검사장비",    "혈액검사, 심전도, 산소포화도 측정기 등"),
    ("치과장비",        "치과 전용 장비 및 기구"),
    ("한방장비",        "한방 치료 전용 장비"),
    ("안과장비",        "안과 전용 장비"),
    ("피부미용장비",    "피부과·성형외과 레이저, IPL 등"),
    ("수술기구소모품",  "수술용 소모성 기구 및 재료"),
    ("기타",            "위 분류에 해당하지 않는 기타 품목"),
]

for r, (cat, desc) in enumerate(cat_desc, 3):
    bg = "F0F9FF" if r % 2 == 0 else "FFFFFF"
    for ci, val in enumerate([cat, desc], 1):
        c = ws2.cell(row=r, column=ci, value=val)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = border
        c.alignment = Alignment(vertical="center")
    ws2.row_dimensions[r].height = 20

# ── 시트3: 작성 가이드 ──
ws3 = wb.create_sheet("작성 가이드")
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 60

ws3.merge_cells("A1:B1")
ws3["A1"] = "작성 가이드"
ws3["A1"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", start_color=COLOR_HEADER)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

guide = [
    ("항목",        "작성 방법"),
    ("카테고리명",  "드롭다운에서 선택하거나 '카테고리 목록' 시트에서 정확히 복사. 없는 카테고리는 담당자에게 문의."),
    ("품목명",      "장비 종류명. 예: X-Ray (DR System) / 초음파진단기 / ICT 간섭파"),
    ("모델명",      "제조사 공식 모델명. 예: Xvision HF 525R"),
    ("제조사",      "회사 공식 명칭 그대로 입력. 예: GE Healthcare (GE 로 단축 금지)"),
    ("가격(원)",    "공급가 기준 원 단위 숫자만. 예: 15000000\n미확정 또는 문의 필요한 경우 빈칸"),
    ("원산지",      "예: 대한민국 / 미국 / 독일 / 이탈리아"),
    ("인증",        "쉼표로 구분. 예: 식약처 허가, CE, FDA, ISO 13485"),
    ("A/S 기간",    "예: 1년 / 2년 / 2년(부품 5년)"),
    ("보증기간",    "예: 1년 / 설치 후 1년"),
    ("제품소개",    "200자 이내 핵심 특징 요약"),
    ("스펙 항목/값","스펙1항목: '관전압'  스펙1값: '40~150 kVp'\n중요한 것부터 최대 5개"),
    ("특이사항",    "예: 콘덴서 포함 / 미팅시 가격제안 / 단종 예정"),
    ("",            ""),
    ("주의사항",    "1. 노란색 샘플 행(4~6행)은 참고 후 삭제하고 7행부터 입력하세요.\n2. 필수 항목(카테고리명, 품목명, 모델명)은 반드시 입력.\n3. 같은 모델명이 이미 DB에 있으면 자동으로 건너뜁니다.\n4. 완성된 파일을 전달해 주시면 일괄 등록해 드립니다."),
]

for r, (k, v) in enumerate(guide, 2):
    c1 = ws3.cell(row=r, column=1, value=k)
    c2 = ws3.cell(row=r, column=2, value=v)
    is_hdr = r == 2
    is_warn = k == "주의사항"
    bg = "DBEAFE" if is_hdr else ("FFFBEB" if is_warn else ("F8FAFC" if r % 2 == 0 else "FFFFFF"))
    h = 40 if "\n" in v else 22
    ws3.row_dimensions[r].height = h
    for c in [c1, c2]:
        c.font = Font(name="Arial", size=10, bold=(is_hdr or is_warn))
        c.fill = PatternFill("solid", start_color=bg)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)

wb.save("C:/Users/jojo/mediquote/장비_정보_입력양식.xlsx")
print("완료: C:/Users/jojo/mediquote/장비_정보_입력양식.xlsx")
