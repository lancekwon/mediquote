"""
대원메디칼 분개장 (1234.xlsx) → 거래처별 장비매입 정리 엑셀

추출 대상: '외 상 매 입 금'이 대변(row[5])에 잡힌 분개의 상품 매입
출력 컬럼: 일자 / 거래처 / 품목설명 / 모델명 / 공급가액 / 부가세 / 합계(VAT포함)
"""
import sys, re, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\jojo\Desktop\1234.xlsx'
OUT = r'C:\Users\jojo\Desktop\거래처별_장비매입_2025_v2.xlsx'

# ----- 1. 분개장 파싱 -----
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Sheet1']
rows = list(ws.iter_rows(values_only=True))

def normalize_date(v):
    if v is None or v == '':
        return ''
    s = str(v).strip()
    # '20250101' 형식
    if s.isdigit() and len(s) == 8:
        return f'2025-{s[4:6]}-{s[6:8]}'
    # '01/02' 형식
    m = re.match(r'^(\d{1,2})/(\d{1,2})$', s)
    if m:
        return f'2025-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
    return s

def extract_model(desc):
    if not desc: return ''
    m = re.search(r'<([^>]+)>', desc)
    return m.group(1).strip() if m else ''

txs = []  # 매입 거래 리스트
i = 0
while i < len(rows):
    row = rows[i]
    # row[4] = '외 상 매 입 금', row[5] = 금액(>0) → 매입 발생
    col4_clean = str(row[4] or '').replace(' ', '')
    if '외상매입금' in col4_clean and row[5] and isinstance(row[5], (int, float)) and row[5] > 0:
        date = normalize_date(row[0])
        total = int(row[5])
        # 다음 4행 안에서 부가세/공급가/품목 찾기
        vat = supply = item_desc = vendor = None
        j = i + 1
        while j < min(i + 6, len(rows)):
            r = rows[j]
            d3 = str(r[3] or '').replace(' ', '')
            if '부가세대급금' in d3 and r[2]:
                vat = int(r[2])
            elif '상품' in d3 and r[2] and len(d3) <= 4:
                supply = int(r[2])
            elif r[3] and r[4] and not r[2] and not r[5]:
                item_desc = str(r[3]).strip()
                vendor = str(r[4]).strip()
                break
            # 새 분개 시작이면 중단
            if j > i + 1 and r[0] and r[1]:
                break
            j += 1

        # 상품 차변(supply)이 잡힌 분개만 = 진짜 장비/물품 매입
        # 노이즈 제외: 외상매출금 입금 충당, 광고료 등 비상품 결제
        if vendor and item_desc and supply and supply > 0:
            txs.append({
                'date': date,
                'vendor': vendor,
                'desc': item_desc,
                'model': extract_model(item_desc),
                'supply': supply or (total - (vat or 0)),
                'vat': vat or 0,
                'total': total,
            })
        i = j
        continue
    i += 1

print(f'추출된 매입 거래: {len(txs)}건')
print(f'거래처 수: {len(set(t["vendor"] for t in txs))}개')

# ----- 2. 거래처별 그룹화 -----
by_vendor = defaultdict(list)
for t in txs:
    by_vendor[t['vendor']].append(t)
# 정렬: 거래처별 합계 내림차순
sorted_vendors = sorted(by_vendor.keys(),
    key=lambda v: -sum(t['total'] for t in by_vendor[v]))

# ----- 3. 엑셀 출력 -----
wb_out = Workbook()

# Sheet 1: 거래처별 요약
ws1 = wb_out.active
ws1.title = '거래처별 요약'
ws1.append(['거래처', '거래 건수', '총 매입금액(VAT포함)', '주요 품목 (최대 3건)'])
for v in sorted_vendors:
    items = by_vendor[v]
    total = sum(t['total'] for t in items)
    top = ' / '.join(t['desc'][:30] for t in items[:3])
    ws1.append([v, len(items), total, top])
ws1.append([])
ws1.append(['합계', sum(len(by_vendor[v]) for v in sorted_vendors),
            sum(t['total'] for t in txs), ''])

# Sheet 2: 거래처별 상세 (모든 거래)
ws2 = wb_out.create_sheet('거래처별 상세')
ws2.append(['거래처', '일자', '품목 설명', '모델명', '공급가액', '부가세', '합계(VAT포함)'])
for v in sorted_vendors:
    items = sorted(by_vendor[v], key=lambda t: t['date'])
    for t in items:
        ws2.append([v, t['date'], t['desc'], t['model'], t['supply'], t['vat'], t['total']])
    # 거래처 소계
    ws2.append([f'  └ {v} 소계', '', '', '', sum(t['supply'] for t in items),
                sum(t['vat'] for t in items), sum(t['total'] for t in items)])
    ws2.append([])

# Sheet 3: 전체 거래 (시간순)
ws3 = wb_out.create_sheet('전체 거래')
ws3.append(['일자', '거래처', '품목 설명', '모델명', '공급가액', '부가세', '합계(VAT포함)'])
for t in sorted(txs, key=lambda x: x['date']):
    ws3.append([t['date'], t['vendor'], t['desc'], t['model'], t['supply'], t['vat'], t['total']])

# Sheet 4: 거래처별 취급 모델
qty_pat = re.compile(r'(\d+)\s*[xX×]\s*[\d,]+')

def extract_qty(desc):
    if not desc: return 0
    m = qty_pat.search(desc)
    return int(m.group(1)) if m else 0

def extract_item_name(desc):
    """모델명 < > 앞부분의 품목명 추출"""
    if not desc: return ''
    # < > 가 있으면 그 앞 부분
    m = re.match(r'^(.+?)\s*<', desc)
    if m: return m.group(1).strip()
    # 없으면 앞부분 (수량 패턴 전까지)
    m = re.match(r'^(.+?)\s+\d+\s*[xX×]', desc)
    if m: return m.group(1).strip()
    # 그 외 전체
    return desc.replace('외', '').strip()

# 거래처 → 모델 → 데이터 집계
from collections import defaultdict
model_data = defaultdict(lambda: defaultdict(lambda: {'count': 0, 'qty': 0, 'amount': 0, 'item_names': set(), 'first_date': '', 'last_date': ''}))
for t in txs:
    v = t['vendor']
    model = t['model'] or '(모델명 미기재)'
    item_name = extract_item_name(t['desc'])
    qty = extract_qty(t['desc'])
    d = model_data[v][model]
    d['count'] += 1
    d['qty'] += qty
    d['amount'] += t['total']
    if item_name: d['item_names'].add(item_name)
    if not d['first_date'] or t['date'] < d['first_date']: d['first_date'] = t['date']
    if not d['last_date'] or t['date'] > d['last_date']: d['last_date'] = t['date']

ws4 = wb_out.create_sheet('거래처별 취급 모델')
ws4.append(['거래처', '품목명', '모델명', '거래 횟수', '총 수량', '총 매입금액(VAT포함)', '최초 거래', '최근 거래'])

for v in sorted_vendors:
    models = model_data[v]
    # 거래처 내 모델별 매입금액 내림차순
    sorted_models = sorted(models.items(), key=lambda x: -x[1]['amount'])
    for model, d in sorted_models:
        item_names = ', '.join(sorted(d['item_names'])) if d['item_names'] else ''
        ws4.append([v, item_names, model, d['count'], d['qty'] or '', d['amount'],
                   d['first_date'], d['last_date']])
    # 거래처 소계
    ws4.append([f'  └ {v} 소계 ({len(sorted_models)}개 모델)', '', '',
                sum(d['count'] for _, d in sorted_models),
                sum(d['qty'] for _, d in sorted_models),
                sum(d['amount'] for _, d in sorted_models), '', ''])
    ws4.append([])

# ----- 4. 서식 -----
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1E40AF')
total_fill = PatternFill('solid', fgColor='F1F5F9')
border = Border(left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1'))

for ws_obj in [ws1, ws2, ws3, ws4]:
    for cell in ws_obj[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 컬럼 폭 조정
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 60

for col, w in zip('ABCDEFG', [22, 12, 50, 22, 16, 14, 18]):
    ws2.column_dimensions[col].width = w
    ws3.column_dimensions[col].width = w

for col, w in zip('ABCDEFGH', [25, 30, 25, 10, 10, 20, 12, 12]):
    ws4.column_dimensions[col].width = w

# 숫자 포맷 (천단위)
for ws_obj, num_cols in [(ws1, ['B', 'C']),
                          (ws2, ['E', 'F', 'G']),
                          (ws3, ['E', 'F', 'G']),
                          (ws4, ['D', 'E', 'F'])]:
    for col in num_cols:
        for cell in ws_obj[col][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

# 소계 행 강조 (Sheet 2 / Sheet 4)
for ws_obj in [ws2, ws4]:
    for row in ws_obj.iter_rows(min_row=2):
        cell = row[0]
        if cell.value and '소계' in str(cell.value):
            for c in row:
                c.fill = total_fill
                c.font = Font(bold=True)

wb_out.save(OUT)
print(f'\n저장 완료: {OUT}')
print(f'시트: 거래처별 요약 / 거래처별 상세 / 전체 거래')
