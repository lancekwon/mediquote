import urllib.request
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPABASE_URL = 'https://dmqzixpappullrnyospj.supabase.co'
SUPABASE_KEY = 'sb_publishable_UcusGk4UNMVEp82y_2jSdA_dQGjd1bH'

def fetch_equipment():
    url = f'{SUPABASE_URL}/rest/v1/equipment?select=cat_name,item_name,model_name,manufacturer,price&order=cat_name.asc,item_name.asc&limit=2000'
    req = urllib.request.Request(url, headers={
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}'
    })
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode())

data = fetch_equipment()
print(f'총 {len(data)}개 장비 데이터 불러옴')

wb = Workbook()
ws = wb.active
ws.title = '장비목록'

# 헤더 스타일
header_fill = PatternFill('solid', start_color='1E293B', end_color='1E293B')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_align = Alignment(horizontal='center', vertical='center')

thin = Side(style='thin', color='CBD5E1')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['카테고리', '품목', '모델명', '제조사', '단가(원)']
col_widths = [20, 28, 30, 20, 16]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 28

# 데이터
row_fill_even = PatternFill('solid', start_color='F8FAFC', end_color='F8FAFC')
data_font = Font(name='Arial', size=10)
data_font_price = Font(name='Arial', size=10)

for row_idx, item in enumerate(data, 2):
    fill = row_fill_even if row_idx % 2 == 0 else None
    vals = [
        item.get('cat_name') or '',
        item.get('item_name') or '',
        item.get('model_name') or '',
        item.get('manufacturer') or '',
        item.get('price'),
    ]
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical='center')
        if fill:
            cell.fill = fill
        if col_idx == 5 and val is not None:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')

# 빈 매핑 컬럼 추가 (세금계산서 품명 매핑용)
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 20

mapping_header = ws.cell(row=1, column=6, value='[매핑] 세금계산서 품명')
mapping_header.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
mapping_header.fill = PatternFill('solid', start_color='7C3AED', end_color='7C3AED')
mapping_header.alignment = header_align
mapping_header.border = border

note_header = ws.cell(row=1, column=7, value='비고')
note_header.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
note_header.fill = PatternFill('solid', start_color='7C3AED', end_color='7C3AED')
note_header.alignment = header_align
note_header.border = border

for row_idx in range(2, len(data) + 2):
    for col_idx in [6, 7]:
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        fill = row_fill_even if row_idx % 2 == 0 else None
        if fill:
            cell.fill = fill

# 틀 고정
ws.freeze_panes = 'A2'

# 자동 필터
ws.auto_filter.ref = f'A1:G{len(data)+1}'

out_path = os.path.join(os.path.dirname(__file__), '장비목록_DB현황.xlsx')
wb.save(out_path)
print(f'저장 완료: {out_path}')
