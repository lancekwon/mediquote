"""
대원메디칼 은행 입출금 내역 → 검토용 매칭 엑셀 생성 (방법 A)

입력: Desktop/0529 대원 입출금내역(수정).xlsx
출력: Desktop/입출금_매입매출반영_검토.xlsx

각 거래에 대해:
- 거래처 자동 매칭 후보 채움
- 유형 자동 추정 (지급/운영비/수금/선급금)
- 사용자가 미매칭/오매칭 건을 직접 수정 (드롭다운 제공)
확정 후 bank_import_apply.py로 매입매출관리에 반영
"""
import sys, re, json, urllib.request, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\jojo\Desktop\0529 대원 입출금내역(수정).xlsx'
OUT = r'C:\Users\jojo\Desktop\입출금_매입매출반영_검토.xlsx'
SB_URL = 'https://nbgubiywavozgigiwkpr.supabase.co'
SB_KEY = 'sb_publishable_L4FVvZBPaNF9BQtoadoPRw_3HeNlPRL'

# ----- 1. 은행 내역 읽기 -----
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
txs = []
for r in rows[7:]:
    if r[0] is None: continue
    txs.append({
        'no': r[0], 'dt': str(r[1] or ''), 'name': r[2] or '',
        'out': int(r[3] or 0), 'in': int(r[4] or 0), 'bal': int(r[5] or 0),
        'memo': r[7] or '', 'jeokyo': r[8] or '',
    })

# ----- 2. 거래처 마스터 -----
req = urllib.request.Request(f'{SB_URL}/rest/v1/manufacturers?select=id,name,vendor_code,aliases',
    headers={'apikey': SB_KEY, 'Authorization': f'Bearer {SB_KEY}'})
mfrs = json.loads(urllib.request.urlopen(req).read())
# 별칭(계좌주명/약칭) 전처리: 쉼표 구분 → 정규화 목록
for m in mfrs:
    m['_aliases'] = [a.strip() for a in (m.get('aliases') or '').split(',') if a.strip()]

BANKS = ['우리','신한','기업','하나','농협','국민','스타뱅','카카오','케이','새마을','수협','부산','대구','광주','전북','경남','SC','씨티','한화생','삼성']
def norm(s):
    return (s or '').replace('　',' ').replace('(주)','').replace('（주）','').replace('주식회사','').replace(' ','').strip()

def match_vendor(name, memo):
    cands = []
    paren = re.findall(r'[(（]([^)）]+)[)）]', name)
    cands += paren
    if memo and '잔액' in memo:
        cands.append(memo.replace('잔액','').strip())
    nm = name
    for b in BANKS:
        if nm.startswith(b): nm = nm[len(b):]; break
    cands.append(nm)
    for c in cands:
        cc = norm(c)
        if len(cc) < 2: continue
        for m in mfrs:
            mn = norm(m['name'])
            if cc and (cc in mn or mn in cc):
                return m['name']
            # 별칭(계좌주명/약칭) 대조 — 사람이름·약칭으로 찍히는 통장 표시 매칭률↑
            for al in m['_aliases']:
                an = norm(al)
                if len(an) < 2: continue
                if cc == an or cc in an or an in cc:
                    return m['name']
    return ''

OPEX_KW = ['부가세','세금','국세','지방세','세무','공과','전기','통신','KT','SKT','ＳＫＴ','보험','카드','수수료','이자','대출','임대','월세','급여','상여','광고','renta','렌탈','입출통지','기장','업무비']
def guess_type(t, vendor):
    blob = f"{t['name']} {t['memo']} {t['jeokyo']}"
    if t['in'] > 0:
        return '수금'
    if '선급' in blob:
        return '선급금'
    if any(k in blob for k in OPEX_KW):
        return '운영비'
    if vendor:
        return '지급'
    return ''  # 미정

# ----- 3. 검토 엑셀 작성 -----
out = Workbook()
sh = out.active
sh.title = '입출금 검토'
headers = ['No','거래일시','송금명(은행표시)','출금액','입금액','통장잔액','메모','적요',
           '자동매칭 거래처','✏️확정 거래처','✏️유형','반영']
sh.append(headers)

auto_matched = 0
for t in txs:
    v = match_vendor(t['name'], t['memo'])
    if v and t['out'] > 0: auto_matched += 1
    typ = guess_type(t, v)
    sh.append([
        t['no'], t['dt'], t['name'], t['out'] or '', t['in'] or '', t['bal'],
        t['memo'], t['jeokyo'], v, v, typ, 'Y'
    ])

# ----- 4. 서식 + 드롭다운 -----
hdr_font = Font(bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='1E40AF')
edit_fill = PatternFill('solid', fgColor='FEF9C3')  # 노란색 — 편집 컬럼
for cell in sh[1]:
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

widths = [5, 20, 28, 14, 14, 14, 22, 14, 22, 22, 12, 7]
for col, w in zip('ABCDEFGHIJKL', widths):
    sh.column_dimensions[col].width = w

# 편집 컬럼(J=확정거래처, K=유형) 노란 배경
for row in range(2, sh.max_row + 1):
    sh.cell(row=row, column=10).fill = edit_fill
    sh.cell(row=row, column=11).fill = edit_fill
    for c in (4, 5, 6):
        sh.cell(row=row, column=c).number_format = '#,##0'

# 유형 드롭다운
dv_type = DataValidation(type='list', formula1='"지급,운영비,수금,선급금,제외"', allow_blank=True)
dv_type.error = '지급/운영비/수금/선급금/제외 중 선택'
dv_type.prompt = '지급=외상차감, 운영비=통장만, 수금=입금, 선급금=별도, 제외=반영안함'
sh.add_data_validation(dv_type)
dv_type.add(f'K2:K{sh.max_row}')

# 반영 드롭다운
dv_yn = DataValidation(type='list', formula1='"Y,N"', allow_blank=False)
sh.add_data_validation(dv_yn)
dv_yn.add(f'L2:L{sh.max_row}')

# 거래처 목록을 별도 시트에 (참고용 — 확정 거래처 채울 때 복사용)
ws_ref = out.create_sheet('거래처목록(참고)')
ws_ref.append(['거래처명', '코드'])
for m in sorted(mfrs, key=lambda x: x['name']):
    ws_ref.append([m['name'], m.get('vendor_code') or ''])
ws_ref.column_dimensions['A'].width = 30
ws_ref.column_dimensions['B'].width = 12
for cell in ws_ref[1]:
    cell.font = hdr_font; cell.fill = hdr_fill

# 안내 시트
ws_help = out.create_sheet('작성안내', 0)
guide = [
    ['■ 입출금 → 매입매출 반영 검토 시트'],
    [''],
    ['1. "입출금 검토" 탭의 노란 칸 2개만 확인/수정하세요:'],
    ['   · ✏️확정 거래처 : 자동매칭이 비었거나 틀린 행만 직접 입력 (거래처목록 탭 참고)'],
    ['   · ✏️유형 : 드롭다운에서 선택'],
    [''],
    ['2. 유형 의미:'],
    ['   · 지급   = 거래처 외상매입금에서 차감 (payment) + 통장 기록'],
    ['   · 운영비 = 세금/통신/카드 등. 통장 잔액에만 기록 (외상 무관)'],
    ['   · 수금   = 입금. 통장 잔액에만 기록'],
    ['   · 선급금 = 병원 선급금 등. 통장에만 기록 (검토 후 별도 처리)'],
    ['   · 제외   = 반영 안 함'],
    [''],
    ['3. 다 채우면 이 파일을 그대로 두고 "다 됐어" 라고 알려주세요.'],
    ['   → 매입매출관리에 일괄 반영합니다 (지급은 거래처 잔액 자동 차감).'],
    [''],
    ['※ "지급"인데 확정 거래처가 비어있으면 반영 시 건너뜁니다.'],
]
for g in guide:
    ws_help.append(g)
ws_help['A1'].font = Font(bold=True, size=13)
ws_help.column_dimensions['A'].width = 80

out.save(OUT)
print(f'저장: {OUT}')
print(f'총 {len(txs)}건 (출금 {sum(1 for t in txs if t["out"]>0)} / 입금 {sum(1 for t in txs if t["in"]>0)})')
print(f'거래처 자동매칭 (출금): {auto_matched}건')
