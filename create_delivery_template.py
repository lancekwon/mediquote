from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
ws = wb.active
ws.title = '납품이력_입력'

thin = Side(style='thin', color='CBD5E1')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style='medium', color='94A3B8'))

# ── 안내 (1행) ──
ws.merge_cells('A1:N1')
notice = ws.cell(row=1, column=1,
    value='✅ 세금계산서 데이터를 그대로 붙여넣으세요. 병원명은 DB에 등록된 이름과 정확히 일치해야 합니다. 같은 납품 건은 병원명+납품일+문서번호가 동일해야 합니다.')
notice.font = Font(name='Arial', size=9, color='92400E')
notice.fill = PatternFill('solid', start_color='FEF3C7', end_color='FEF3C7')
notice.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 24

# ── DB 컬럼명 (2행) ──
db_cols = [
    # deliveries 테이블
    'hospitals.name', 'delivered_date', 'doc_no', 'supply_amount', 'vat', 'total_amount', 'notes',
    # delivery_items 테이블
    'item_code', 'item_name', 'model_name', 'spec', 'unit', 'quantity', 'unit_price',
]
for i, col in enumerate(db_cols, 1):
    cell = ws.cell(row=2, column=i, value=col)
    cell.font = Font(name='Arial', size=8, color='94A3B8', italic=True)
    cell.fill = PatternFill('solid', start_color='F1F5F9', end_color='F1F5F9')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
ws.row_dimensions[2].height = 14

# ── 테이블 구분 헤더 (3행) ──
ws.merge_cells('A3:G3')
delivery_group = ws.cell(row=3, column=1, value='📦 납품 건 정보 (deliveries)')
delivery_group.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
delivery_group.fill = PatternFill('solid', start_color='B45309', end_color='B45309')
delivery_group.alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('H3:N3')
item_group = ws.cell(row=3, column=8, value='📋 품목 정보 (delivery_items)')
item_group.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
item_group.fill = PatternFill('solid', start_color='1E3A5F', end_color='1E3A5F')
item_group.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[3].height = 20

# ── 헤더 (4행) ──
headers = [
    # deliveries
    ('병원명 *', True, 22),
    ('납품일 *', True, 14),
    ('문서번호', False, 14),
    ('공급가액', False, 14),
    ('부가세', False, 12),
    ('합계금액', False, 14),
    ('메모', False, 20),
    # delivery_items
    ('품목코드', False, 14),
    ('품목명 *', True, 22),
    ('모델명', False, 20),
    ('규격', False, 16),
    ('단위', False, 8),
    ('수량', False, 8),
    ('단가', False, 14),
]

for col_idx, (label, required, width) in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=label)
    if col_idx <= 7:
        bg = 'D97706' if required else 'F59E0B'
    else:
        bg = '1E3A5F' if required else '334155'
    cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thick_bottom
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[4].height = 26

# ── 샘플 데이터 (5~9행: 센텀메디의원 예시) ──
samples = [
    ['센텀메디의원', '2023-08-28', '10174', 400000, 40000, 440000, '',   '11005-1081', 'ExaminingTable', 'KT-108,온열', '', '', 2, 220000],
    ['센텀메디의원', '2023-08-28', '10174', 309091, 30909, 340000, '',   '11009-905',  'ExaminingTableOp.', 'Siderail', '', '', 2, 170000],
    ['센텀메디의원', '2023-08-28', '10174', 65455,  6545,  72000,  '',   '11009-908',  'ExaminingTableOp.', 'IVPole,T자링거대', '', '', 2, 36000],
    ['센텀메디의원', '2023-08-28', '10174', 25455,  2545,  28000,  '',   '11009-910',  'ExaminingTableOp.', '링겔꽂이', '', '', 2, 14000],
    ['센텀메디의원', '2023-08-28', '10174', 80000,  8000,  88000,  '',   '79000-555',  'Carriage Charges', '', '', '', 2, 44000],
]
for row_idx, row_data in enumerate(samples, 5):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=9, color='0369A1', italic=True)
        cell.fill = PatternFill('solid', start_color='EFF6FF', end_color='EFF6FF')
        cell.alignment = Alignment(vertical='center')
        cell.border = border
        if col_idx in [4,5,6,14] and isinstance(val, (int,float)):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[row_idx].height = 18

# ── 입력 영역 (10행~310행) ──
row_fill_even = PatternFill('solid', start_color='FFFBEB', end_color='FFFBEB')
row_fill_item = PatternFill('solid', start_color='F8FAFC', end_color='F8FAFC')
data_font = Font(name='Arial', size=10)
num_fmt_cols = {4, 5, 6, 14}  # 공급가액, 부가세, 합계, 단가

for row_idx in range(10, 310):
    fill = row_fill_even if row_idx % 2 == 0 else None
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical='center')
        if col_idx <= 7 and fill:
            cell.fill = fill
        elif col_idx > 7:
            cell.fill = row_fill_item if row_idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
        if col_idx in num_fmt_cols:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[row_idx].height = 18

ws.freeze_panes = 'A5'
ws.auto_filter.ref = 'A4:N309'

# ── Sheet2: 업로드 가이드 ──
guide = wb.create_sheet('📋 가이드')
guide.column_dimensions['A'].width = 18
guide.column_dimensions['B'].width = 65

guide_rows = [
    ('', ''),
    ('  📌 작성 규칙', ''),
    ('', ''),
    ('  병원명', 'DB 병원관리에 등록된 이름과 완전히 동일하게 입력'),
    ('  같은 납품건', '병원명 + 납품일 + 문서번호가 같으면 같은 deliveries 행으로 묶음'),
    ('  품목 한 행 = 1행', 'delivery_items 테이블에 각각 1행으로 들어감'),
    ('  금액', '공급가액 + 부가세 = 합계금액 (세금계산서 그대로)'),
    ('  납품일 형식', 'YYYY-MM-DD (예: 2023-08-28)'),
    ('', ''),
    ('  ⚠️ 주의사항', ''),
    ('', ''),
    ('  중복 주의', '같은 데이터를 두 번 import하면 중복 생성됨'),
    ('  병원 미등록', '병원관리에 없는 병원명 → import 실패. 먼저 병원 등록 필요'),
    ('  supply_amount 계산', '단가 × 수량 / 1.1 (부가세 별도 시) 또는 세금계산서 금액 그대로'),
    ('', ''),
    ('  📊 DB 연결 구조', ''),
    ('', ''),
    ('  hospitals', '← 병원명으로 hospital_id 조회'),
    ('  deliveries', '납품 건 단위 (병원1개 = 날짜별 여러 건 가능)'),
    ('  delivery_items', '품목 단위 (delivery 1건 = 여러 품목)'),
]

for r, (a, b) in enumerate(guide_rows, 1):
    ca = guide.cell(row=r, column=1, value=a)
    cb = guide.cell(row=r, column=2, value=b)
    ca.font = Font(name='Arial', size=10, bold=('📌' in a or '⚠️' in a or '📊' in a))
    cb.font = Font(name='Arial', size=10)
    if '📌' in a or '⚠️' in a or '📊' in a:
        guide.merge_cells(f'A{r}:B{r}')
        guide.cell(row=r, column=1).value = a
        guide.cell(row=r, column=1).font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        guide.cell(row=r, column=1).fill = PatternFill('solid', start_color='1E293B', end_color='1E293B')
    guide.row_dimensions[r].height = 20

out_path = 'C:/Users/jojo/mediquote/납품이력_입력양식.xlsx'
wb.save(out_path)
print(f'저장 완료: {out_path}')
